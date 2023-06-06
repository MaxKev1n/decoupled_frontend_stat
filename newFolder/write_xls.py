import openpyxl
import numpy as np

def write_xls():
    np.set_printoptions(formatter={'all': lambda x: str(x)}, threshold=100)

    xs_dev_mpki = []
    xs_dev_misRate = []
    xs_dev_LTAGE_mpki = []
    xs_dev_LTAGE_misRate = []
    stream_mpki = []
    stream_misRate = []
    #decoupled_last_mpki = []
    #decoupled_last_misRate = []

    for i in range(0, 3):
        read_file = []
        graph_name = []

        if i == 0:
            read_file = 'xs-dev.txt'
            graph_name = 'xs-dev'
        elif i == 1:
            read_file = 'xs-dev-LTAGE.txt'
            graph_name = 'xs-dev-LTAGE'
        else:
            read_file = 'stream.txt'
            graph_name = 'stream'

        f1 = open('workloadBranch/' + read_file)

        length = int(f1.readline())

        workload_name = []
        mpki = []
        misRate = []

        for line in f1:
            stat = line.split()
            workload_name.append(str(stat[0]))
            mpki.append(float(stat[4]))
            misRate.append(float(stat[5]))

        if i == 0:
            xs_dev_misRate = misRate
            xs_dev_mpki = mpki
        elif i == 1:
            xs_dev_LTAGE_misRate = misRate
            xs_dev_LTAGE_mpki = mpki
        else:
            stream_misRate = misRate
            stream_mpki = mpki

    wb = openpyxl.load_workbook('xls/MPKI.xlsx')
    sheet = wb['Sheet']

    for i in range(length):
        #sheet.append([str(workload_name[i]), str(decoupled_update_mpki[i]), str(decoupled_last_mpki[i]), str(decoupled_mpki[i]), str(coupled_mpki[i])])
        sheet.append([str(workload_name[i]), str(xs_dev_mpki[i]), str(xs_dev_LTAGE_mpki[i]), str(stream_mpki[i])])
    wb.save('xls/MPKI.xlsx')

    wb2 = openpyxl.load_workbook('xls/misRate.xlsx')
    sheet2 = wb2['Sheet']

    for i in range(length):
        #sheet2.append([str(workload_name[i]), str(decoupled_update_misRate[i]), str(decoupled_last_misRate[i]), str(decoupled_misRate[i]), str(coupled_misRate[i])])
        sheet2.append([str(workload_name[i]), str(xs_dev_misRate[i]), str(xs_dev_LTAGE_misRate[i]), str(stream_misRate[i])])
    wb2.save('xls/misRate.xlsx')

def write_xls_stallReason(str):
    np.set_printoptions(formatter={'all': lambda x: str(x)}, threshold=100)
    label = ['', 'NoStall', 'IcacheStall', 'ITlbStall', 'DTlbStall', 'BpStall', 'IntStall', 'TrapStall', 'FragStall',
             'SquashStall', 'FetchBufferInvalid', 'InstMisPred', 'InstSquashed', 'SerializeStall', 'LongExecute',
             'InstNotReady',
             'LoadL1Stall', 'LoadL2Stall', 'LoadL3Stall', 'StoreL1Stall', 'StoreL2Stall', 'StoreL3Stall',
             'ResumeUnblock', 'CommitSquash',
             'Other']

    stage = ''

    for i in range(1):
        if i == 0:
            stage = 'dispatch'
        elif i == 1:
            stage = 'decode'
        elif i == 2:
            stage = 'rename'
        else:
            stage = 'dispatch'

        f1 = open('workloadStallReason/' + str + '-' + stage + '-cpt.txt')
        wb = openpyxl.load_workbook('xls/' + str + '.xlsx')
        sheet = wb['Sheet']
        sheet.append(label)

        for line in f1:
            stat = line.split()
            sheet.append(stat)

        wb.save('xls/' + str + '.xlsx')


def write_xls_ipc():
    np.set_printoptions(formatter={'all': lambda x: str(x)}, threshold=100)

    xs_dev = []
    xs_dev_LTAGE = []
    stream = []

    for i in range(0, 3):
        read_file = []
        graph_name = []

        if i == 0:
            read_file = 'xs-dev.txt'
            graph_name = 'xs-dev'
        elif i == 1:
            read_file = 'xs-dev-LTAGE.txt'
            graph_name = 'xs-dev-LTAGE'
        else:
            read_file = 'stream.txt'
            graph_name = 'stream'

        f1 = open('workloadIPC/' + read_file)

        length = int(f1.readline())

        workload_name = []
        ipc = []

        for line in f1:
            stat = line.split()
            workload_name.append(str(stat[0]))
            ipc.append(float(stat[1]))

        if i == 0:
            xs_dev = ipc
        elif i == 1:
            xs_dev_LTAGE = ipc
        else:
            stream= ipc

    wb = openpyxl.load_workbook('xls/IPC.xlsx')
    sheet = wb['Sheet']

    for i in range(length):
        #sheet.append([str(workload_name[i]), str(decoupled_update_mpki[i]), str(decoupled_last_mpki[i]), str(decoupled_mpki[i]), str(coupled_mpki[i])])
        sheet.append([str(workload_name[i]), str(xs_dev[i]), str(xs_dev_LTAGE[i]), str(stream[i])])
    wb.save('xls/IPC.xlsx')


write_xls_ipc()